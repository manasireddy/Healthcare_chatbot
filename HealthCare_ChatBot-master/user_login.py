import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
import subprocess

def validate_login():
    username = username_entry.get()
    password = password_entry.get()

    try:
        # Load workbook
        wb = load_workbook("registrations.xlsx")
        ws = wb.active
        
        # Iterate through rows and check for matching username and password
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[4] == username and row[5] == password:
                messagebox.showinfo("Success", "Login successful!")
                open_healthcare_app()  # Open the Health_Care_App.py file
                root.withdraw()  # Hide the login window
                return  # Exit the function after successful login
        
        # If no matching credentials found
        messagebox.showerror("Error", "Incorrect username or password.")
    except Exception as e:
        print("Error:", e)
        messagebox.showerror("Error", "Failed to validate login. Please try again.")

def open_healthcare_app():
    try:
        subprocess.Popen(["python", "Health_Care_App.py"])
    except Exception as e:
        print("Error:", e)
        messagebox.showerror("Error", "Failed to open Health_Care_App.py.")

def open_registration_window():
    registration_window = tk.Toplevel(root)
    registration_window.title("Registration Page")
    
    # Add registration widgets with styling
    style = ttk.Style()
    style.configure("TLabel", font=("Arial", 12))
    style.configure("TEntry", font=("Arial", 12))
    style.configure("TButton", font=("Arial", 12))

    name_label = ttk.Label(registration_window, text="Name:")
    name_label.grid(row=0, column=0, padx=10, pady=5, sticky="e")
    name_entry = ttk.Entry(registration_window)
    name_entry.grid(row=0, column=1, padx=10, pady=5)

    email_label = ttk.Label(registration_window, text="Email:")
    email_label.grid(row=1, column=0, padx=10, pady=5, sticky="e")
    email_entry = ttk.Entry(registration_window)
    email_entry.grid(row=1, column=1, padx=10, pady=5)

    gender_label = ttk.Label(registration_window, text="Gender:")
    gender_label.grid(row=2, column=0, padx=10, pady=5, sticky="e")
    gender_entry = ttk.Entry(registration_window)
    gender_entry.grid(row=2, column=1, padx=10, pady=5)

    blood_group_label = ttk.Label(registration_window, text="Blood Group:")
    blood_group_label.grid(row=3, column=0, padx=10, pady=5, sticky="e")
    blood_group_entry = ttk.Entry(registration_window)
    blood_group_entry.grid(row=3, column=1, padx=10, pady=5)

    username_label = ttk.Label(registration_window, text="Username:")
    username_label.grid(row=4, column=0, padx=10, pady=5, sticky="e")
    username_entry_reg = ttk.Entry(registration_window)
    username_entry_reg.grid(row=4, column=1, padx=10, pady=5)

    password_label = ttk.Label(registration_window, text="Password:")
    password_label.grid(row=5, column=0, padx=10, pady=5, sticky="e")
    password_entry_reg = ttk.Entry(registration_window, show="*")
    password_entry_reg.grid(row=5, column=1, padx=10, pady=5)

    register_button = ttk.Button(registration_window, text="Register", command=lambda: register(registration_window, name_entry.get(), email_entry.get(), gender_entry.get(), blood_group_entry.get(), username_entry_reg.get(), password_entry_reg.get()))
    register_button.grid(row=6, columnspan=2, pady=10)

def register(registration_window, name, email, gender, blood_group, username, password):
    # Create or load workbook
    try:
        wb = load_workbook("registrations.xlsx")
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Email", "Gender", "Blood Group", "Username", "Password"])

    # Append registration data to worksheet
    try:
        ws.append([name, email, gender, blood_group, username, password])
        wb.save("registrations.xlsx")
        messagebox.showinfo("Success", "Registration successful!")
        registration_window.destroy()  # Close registration window after successful registration
        check_login_credentials(username, password)
    except Exception as e:
        print("Error:", e)
        messagebox.showerror("Error", "Failed to register. Please try again.")

def check_login_credentials(username, password):
    if username == "admin" and password == "password":
        messagebox.showinfo("Success", "Login successful!")
    else:
        messagebox.showerror("Error", "Incorrect username or password.")

# Create main window
root = tk.Tk()
root.title("Login Page")

# Add styling to login widgets
style = ttk.Style()
style.configure("TLabel", font=("Arial", 12))
style.configure("TEntry", font=("Arial", 12))
style.configure("TButton", font=("Arial", 12))

username_label = ttk.Label(root, text="Username:")
username_label.grid(row=0, column=0, padx=10, pady=5, sticky="e")

username_entry = ttk.Entry(root)
username_entry.grid(row=0, column=1, padx=10, pady=5)

password_label = ttk.Label(root, text="Password:")
password_label.grid(row=1, column=0, padx=10, pady=5, sticky="e")

password_entry = ttk.Entry(root, show="*")
password_entry.grid(row=1, column=1, padx=10, pady=5)

login_button = ttk.Button(root, text="Login", command=validate_login)
login_button.grid(row=2, column=0, columnspan=2, pady=10)

register_button = ttk.Button(root, text="Register", command=open_registration_window)
register_button.grid(row=3, column=0, columnspan=2, pady=5)

root.mainloop()
